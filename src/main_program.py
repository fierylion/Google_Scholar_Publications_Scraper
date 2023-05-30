import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

#create excel sheet and fill in user details
def create_excel(wb, person_details):
    # document of structure {position: lecturer, name:baraka, verification:baraka, fields:[doctor, engineer, cousing], details:[[baraka, baraka, baraka, dafda], [baraka, baraka, baraka], [baraka, baraka, baraka]]}
    details = person_details.get('details') # list of lists
    publication_fields = ['No', 'Title','From', 'Authors', 'Date', 'Source', 'Volume', 'Pages', 'Publisher', 'Citations','Conference' ,'URL']
    ws  = wb.create_sheet(f'Sheet-{person_details.get("name")}', 0)
    ws['A1'] = 'Position'
    ws['B1'] = person_details.get('position')
    ws['A2'] = 'Name'
    ws['B2'] = person_details.get('name')

    ws['A3'] = 'Verification Details'
    ws['B3'] = person_details.get('verification')
    ws['A4'] = 'Fields'
    for (ind, fd) in enumerate(person_details.get('fields')):
        ws.cell(row=4, column=(ind+2)).value = fd

    #fill in details table
    def fill_details_table():
        no_rows = len(details)
        for i in range(no_rows):
            for j in range(len(details[i])):
                ws.cell(row=(i+6), column=(j+1)).value = details[i][j]
    fill_details_table()
    current_row = len(details) + 6
    ws.cell(row=current_row+1, column=1).value = 'Publications'
    for (ind, fd) in enumerate(publication_fields):
        ws.cell(row=(current_row+2), column=(ind+1)).value = fd
    return (ws, current_row + 3)

#write single publication to excel
def publication_details_excel(ws, current_row, pb):
    # Expected document structure: {title: title, authors: authors, from_: from,  date: date, source: source, volume: volume, pages: pages, publisher: publisher, citations: citations, url: url}
    ws.cell(row=current_row, column=1).value = pb.get('no')
    ws.cell(row=current_row, column=2).value = pb.get('title', '')
    ws.cell(row=current_row, column=3).value = pb.get('authors', '')
    ws.cell(row=current_row, column=4).value = pb.get('from_', '')
    ws.cell(row=current_row, column=5).value = pb.get('publication date', '')
    ws.cell(row=current_row, column=6).value = pb.get('source', '')
    ws.cell(row=current_row, column=7).value = pb.get('volume', '')
    ws.cell(row=current_row, column=8).value = pb.get('pages', '')
    ws.cell(row=current_row, column=9).value = pb.get('publisher', '')
    ws.cell(row=current_row, column=10).value = pb.get('total citations', '')
    ws.cell(row=current_row, column=11).value = pb.get('conference', '')
    ws.cell(row=current_row, column=12).value = pb.get('url', '')
    return current_row + 1


def fetch_user_details(dr):
    user_details = {}
    def details_table():
        table = dr.find_element(By.ID, 'gsc_rsb_st')
        rows = table.find_elements(By.TAG_NAME, 'tr')
        details = []
        for (ind, row) in enumerate(rows):
            th_or_tr = 'th' if ind == 0 else 'td'
            details.append([cell.text for cell in row.find_elements(By.TAG_NAME, th_or_tr)])
        return details
    try:
        basic_info = dr.find_element(By.ID, 'gsc_prf_i')
        user_details['name'] = basic_info.find_element(By.ID, 'gsc_prf_inw').text
        user_details['position'] = basic_info.find_element(By.CSS_SELECTOR, '.gsc_prf_il').text.split(',')[0]
        user_details['verification'] = basic_info.find_element(By.ID, 'gsc_prf_ivh').text
        user_details['fields'] = [fd.text for fd in basic_info.find_elements(By.CSS_SELECTOR, '#gsc_prf_int>a')]
        user_details['details'] = details_table()
        return user_details
    except Exception as e:
        print(e)
def fetch_publications(dr,link, no):
    publications = {}
    time.sleep(2)
    def check_presence_of_from():
        try:
            from_ = dr.find_element(By.ID, 'gsc_oci_title_gg').text
            return from_
        except Exception as e:
            print(e)
            return False

    try:
        publications['no'] = no
        publications['url'] = link
        pb= dr.find_element(By.ID, "gsc_oci_title_wrapper")
        from_ = check_presence_of_from()
        publications['from_'] = from_ if from_ else ''
        publications['title'] = pb.find_element(By.ID, 'gsc_oci_title').text
        main_details = dr.find_elements(By.CSS_SELECTOR, '#gsc_oci_table>div')
        main_details.pop()
        for detail in main_details:
            key = detail.find_element(By.CSS_SELECTOR, '.gsc_oci_field').text.lower()
            value = detail.find_element(By.CSS_SELECTOR, '.gsc_oci_value').text if key!='total citations' else detail.find_element(By.CSS_SELECTOR, '.gsc_oci_value a').text.split(' ')[2]
            publications[key] = value
        dr.back()
        return publications
    except Exception as e:
        print(e)


def main(url):
    driver = webdriver.Chrome()
    driver.get(url)
    driver.implicitly_wait(2)
    user_details = {}
    try:
        user_details = fetch_user_details(driver)
        wb = openpyxl.Workbook()
        (ws, current_row) = create_excel(wb, user_details)
        publications = driver.find_elements(By.CSS_SELECTOR, '#gsc_a_b>tr')
        for (ind, pb) in enumerate(publications):
            page_link =pb.find_element(By.TAG_NAME, 'a')
            link = page_link.get_attribute('href')
            page_link.click()
            publication_details = fetch_publications(driver,link,  ind+1)
            current_row = publication_details_excel(ws, current_row, publication_details)
            print(f'Publication {ind+1} done')
            if(ind == len(publications)-1):
                try:
                    show_more = driver.find_element(By.ID, 'gsc_bpf_more')
                    show_more.click()
                    time.sleep(2)
                    new_publications = driver.find_elements(By.CSS_SELECTOR, '#gsc_a_b>tr')
                    # remove repetitions
                    for pb in new_publications:
                        if pb in publications:
                            continue
                        else:
                            publications.append(pb)
                except Exception as e:
                    print(e)
                    break
        wb.save('results/output.xlsx')

    except Exception as e:
        print(e)

main("https://scholar.google.com/citations?view_op=list_works&hl=en&hl=en&user=kUE4mw4AAAAJ")
